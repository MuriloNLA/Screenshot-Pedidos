# exportarpedidos.py
# Le PDFs de pedidos e adiciona os dados extraidos em uma planilha Excel (.xlsx).
# Ignora PDFs que ja foram processados (verificacao pelo nome do arquivo).
#
# Dependencias:
#     pip install pdfplumber openpyxl
#
# Uso no terminal:
#     python exportarpedidos.py <pdf_ou_pasta> <planilha.xlsx>

import re
import sys
import os
import pdfplumber
import openpyxl
from pathlib import Path


# ---------------------------------------------------------------------------
# COLUNAS
# ---------------------------------------------------------------------------
COLUNAS = ["ID", "Data", "Hora", "Tipo", "Cliente", "Frete", "Vendedor", "QuantItem", "TempoLevado"]


# ---------------------------------------------------------------------------
# LIMPEZA DE TEXTO
# ---------------------------------------------------------------------------

def limpar_nome(texto):
    """Remove lixo apos palavras-chave ou sequencias de numeros/simbolos."""
    if not texto:
        return ""
    lixo = re.split(
        r'\s+(IE|CNPJ|CPF|TELEFONE|FONE|TEL|CEP|INSC|ISENTO|:)\s*[:\d]',
        texto, flags=re.IGNORECASE
    )
    nome = lixo[0].strip()
    nome = re.sub(r'[\s:]+[\d./-]+$', '', nome).strip()
    nome = re.sub(r'\s{2,}', ' ', nome)
    return nome.strip()


def title_case(texto):
    """Converte para Title Case (primeira letra de cada palavra maiuscula)."""
    if not texto:
        return ""
    return texto.strip().title()


# ---------------------------------------------------------------------------
# EXTRACAO DO PDF
# ---------------------------------------------------------------------------

def extrair_texto_pdf(caminho_pdf):
    texto = ""
    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto += pagina.extract_text(layout=True) or ""
            texto += "\n"
    return texto


def buscar(padrao, texto, grupo=1):
    match = re.search(padrao, texto, re.IGNORECASE)
    return match.group(grupo).strip() if match else ""


def extrair_dados(texto):
    dados = {}

    dados["data_raw"] = buscar(r"DATA DO PEDIDO\s*:\s*(\d{2}/\d{2}/\d{4})", texto)
    dados["hora_raw"] = buscar(r"DATA DO PEDIDO\s*:\s*\d{2}/\d{2}/\d{4}\s+(\d{2}:\d{2})", texto)

    tipo_raw = buscar(r"TIPO\s+PEDIDO\s*:\s*(.+)", texto)
    dados["tipo"] = title_case(tipo_raw)

    cliente_raw = buscar(r"NOME\s*:([^\n]+)", texto)
    cliente_raw = re.split(r"\s{3,}|\t", cliente_raw)[0].strip()
    dados["cliente"] = limpar_nome(cliente_raw)

    frete_conta = buscar(r"FRETE POR CONTA DO CLIENTE\s*:\s*(\S+)", texto)
    if frete_conta.upper() in ("NAO", "NAO", ""):
        dados["frete"] = "CIF"
    else:
        dados["frete"] = "FOB"

    vendedor_raw = buscar(r"REPRESENTANTE\s*:([^\n]+)", texto)
    vendedor_raw = re.split(r"\s{3,}|\t", vendedor_raw)[0].strip()
    dados["vendedor"] = limpar_nome(vendedor_raw)

    itens = re.findall(r"^\s*\d+\s+[A-Z]", texto, re.MULTILINE)
    dados["quant_item"] = len(itens) if itens else 1

    dados["tempo_levado"] = ""

    return dados


# ---------------------------------------------------------------------------
# PLANILHA
# ---------------------------------------------------------------------------

def obter_proximo_id(ws):
    ultimo = 0
    for linha in ws.iter_rows(min_row=2, values_only=True):
        if linha[0] and str(linha[0]).isdigit():
            ultimo = max(ultimo, int(linha[0]))
    return ultimo + 1


def abrir_ou_criar_planilha(caminho):
    if os.path.exists(caminho):
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        for col, nome in enumerate(COLUNAS, start=1):
            ws.cell(row=1, column=col, value=nome)

    cabecalho = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if "Arquivo" not in cabecalho:
        ws.cell(row=1, column=ws.max_column + 1, value="Arquivo")

    return wb, ws


def get_arquivos_ja_processados(ws):
    processados = set()
    col_arquivo = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "Arquivo":
            col_arquivo = c
            break
    if col_arquivo is None:
        return processados
    for linha in ws.iter_rows(min_row=2, values_only=True):
        if len(linha) >= col_arquivo and linha[col_arquivo - 1]:
            processados.add(str(linha[col_arquivo - 1]).strip())
    return processados


def adicionar_linha(ws, novo_id, dados, nome_arquivo):
    ws.append([
        novo_id,
        dados.get("data_raw", ""),
        dados.get("hora_raw", ""),
        dados.get("tipo", ""),
        dados.get("cliente", ""),
        dados.get("frete", ""),
        dados.get("vendedor", ""),
        dados.get("quant_item", ""),
        dados.get("tempo_levado", ""),
        nome_arquivo
    ])


# ---------------------------------------------------------------------------
# PROCESSAMENTO
# ---------------------------------------------------------------------------

def processar_pdf(caminho_pdf, caminho_xlsx):
    nome_arquivo = Path(caminho_pdf).name

    wb, ws = abrir_ou_criar_planilha(caminho_xlsx)

    ja_processados = get_arquivos_ja_processados(ws)
    if nome_arquivo in ja_processados:
        print(f"  [IGNORADO] '{nome_arquivo}' ja esta na planilha.")
        return

    print(f"  [NOVO] Processando: {nome_arquivo}")
    texto = extrair_texto_pdf(caminho_pdf)
    dados = extrair_dados(texto)

    novo_id = obter_proximo_id(ws)
    adicionar_linha(ws, novo_id, dados, nome_arquivo)
    wb.save(caminho_xlsx)

    print(f"  [OK] ID {novo_id} | {dados.get('cliente')} | {nome_arquivo}")


def main():
    if len(sys.argv) < 3:
        print("Uso: python exportarpedidos.py <pdf_ou_pasta> <planilha.xlsx>")
        return

    entrada  = sys.argv[1]
    planilha = sys.argv[2]

    if os.path.isdir(entrada):
        pdfs = list(Path(entrada).glob("*.pdf"))
    else:
        pdfs = [Path(entrada)]

    print(f"\nProcessando {len(pdfs)} arquivo(s)...\n")

    for pdf in pdfs:
        try:
            processar_pdf(str(pdf), planilha)
        except Exception as e:
            print(f"Erro em {pdf}: {e}")

    print("\nFinalizado!")

if __name__ == "__main__":
    main()
